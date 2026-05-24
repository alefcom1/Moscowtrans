#!/usr/bin/env python3
"""
Экспорт вопросов RTAP из JSON в Excel для согласования с заказчиком.
Использование: python3 export-to-excel.py
Создаёт: questions-for-review.xlsx
"""

import json
import glob
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    raise

COLUMNS = [
    ('A', 'id',          6,  'ID'),
    ('B', 'topic',       12, 'Тематика'),
    ('C', 'level',       14, 'Уровень'),
    ('D', 'type',        6,  'Тип'),
    ('E', 'lang',        6,  'Язык'),
    ('F', 'question',    40, 'Вопрос / инструкция'),
    ('G', 'source',      40, 'Source EN (для BT/FE/RO)'),
    ('H', 'option_1',    30, 'Вариант 1'),
    ('I', 'option_2',    30, 'Вариант 2'),
    ('J', 'option_3',    30, 'Вариант 3'),
    ('K', 'option_4',    30, 'Вариант 4'),
    ('L', 'correct',     8,  'Правильный (1-4)'),
    ('M', 'explanation', 40, 'Объяснение'),
    ('N', 'difficulty',  10, 'Сложность (1-5)'),
    ('O', 'STATUS',      12, 'СТАТУС'),
    ('P', 'COMMENT',     30, 'КОММЕНТАРИЙ'),
]

HEADER_FILL   = PatternFill('solid', fgColor='1A3C6E')
HEADER_FONT   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
DATA_FONT     = Font(name='Calibri', size=10)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical='top')
STATUS_COLORS = {'ок': 'C6EFCE', 'правка': 'FFEB9C', 'удалить': 'FFC7CE'}

def flatten(q, idx):
    payload = q.get('payload', {})
    options = payload.get('options', [])
    correct_idx = payload.get('correct', 0)

    # For tm/ro types: encode correct as JSON string
    if q['type'] in ('tm', 'ro'):
        correct_str = json.dumps(payload.get('pairs' if q['type'] == 'tm' else 'correct_order'), ensure_ascii=False)
    else:
        correct_str = str(correct_idx + 1) if isinstance(correct_idx, int) else str(correct_idx)

    return {
        'id':          '',
        'topic':       q.get('topic', ''),
        'level':       q.get('level', ''),
        'type':        q.get('type', ''),
        'lang':        q.get('lang', 'en'),
        'question':    q.get('question', ''),
        'source':      payload.get('source', ''),
        'option_1':    options[0] if len(options) > 0 else '',
        'option_2':    options[1] if len(options) > 1 else '',
        'option_3':    options[2] if len(options) > 2 else '',
        'option_4':    options[3] if len(options) > 3 else '',
        'correct':     correct_str,
        'explanation': q.get('explanation', ''),
        'difficulty':  str(q.get('difficulty', 3)),
        'STATUS':      'ок',
        'COMMENT':     '',
    }

def style_header(ws):
    for col_letter, key, width, label in COLUMNS:
        cell = ws[f'{col_letter}1']
        cell.value = label
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

def write_rows(ws, rows):
    for i, row_data in enumerate(rows, start=2):
        for col_letter, key, _, _ in COLUMNS:
            cell = ws[f'{col_letter}{i}']
            cell.value     = row_data.get(key, '')
            cell.font      = DATA_FONT
            cell.alignment = WRAP_ALIGN
            ws.row_dimensions[i].height = 60

        # Color STATUS column
        status_cell = ws[f'O{i}']
        val = status_cell.value or 'ок'
        fill_color = STATUS_COLORS.get(val, 'FFFFFF')
        status_cell.fill = PatternFill('solid', fgColor=fill_color)

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    topics = ['technical', 'legal', 'medical', 'it']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total = 0
    for topic in topics:
        json_path = os.path.join(script_dir, f'questions-{topic}.json')
        if not os.path.exists(json_path):
            print(f'⚠️  Файл не найден: {json_path}')
            continue

        with open(json_path, encoding='utf-8') as f:
            questions = json.load(f)

        ws = wb.create_sheet(title=topic.capitalize())
        style_header(ws)

        rows = [flatten(q, i) for i, q in enumerate(questions)]
        write_rows(ws, rows)

        ws.auto_filter.ref = f'A1:P{len(rows) + 1}'
        total += len(rows)
        print(f'✅ {topic}: {len(rows)} вопросов')

    out_path = os.path.join(script_dir, 'questions-for-review.xlsx')
    wb.save(out_path)
    print(f'\n📊 Excel сохранён: {out_path}')
    print(f'📝 Итого вопросов: {total}')

if __name__ == '__main__':
    main()
